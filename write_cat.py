import requests
import openpyxl
import os
import configparser
import re
import json

class CKANManager:
    '''Interacts with the CKAN instance'''
    def __init__(self, api_url, api_key):
        self.api_url = api_url
        self.api_key = api_key
        self.headers = {
            'Authorization': self.api_key,
            'Content-Type': 'application/json'
        }

    def get(self, endpoint, params=None):
        url = f"{self.api_url}{endpoint}"
        response = requests.get(url, params=params, headers=self.headers)
        return self._handle_response(response)
    
    def post(self, endpoint, data=None):
        url = f"{self.api_url}{endpoint}"
        response = requests.post(url, json=data, headers=self.headers)
        return self._handle_response(response)

    def _handle_response(self, response):
        if response.status_code in [200, 201]:
            return response.json()
        else:
            raise Exception(f"API Error: {response.status_code} - {response.text}")

    def fetch_package_data(self, catalog_name):
        '''Fetches catalog data by its name'''
        endpoint = f"/api/3/action/package_show?id={catalog_name}"
        response = self.get(endpoint)
        
        if not response.get('success', False):
            raise Exception(f"Failed to fetch catalog: {response.get('error', 'Unknown error')}")
        
        return response

class ExcelHandler:
    '''Handles interaction with the Excel file'''
    def __init__(self, path):
        self.path = path
        try:
            self.wb = openpyxl.load_workbook(self.path)
        except Exception as e:
            print(f"Error: {str(e)}")
            return
        self.ws = self.wb.worksheets[0]  

    # def check_hauptkategorien(self, fetched_groups):
    # does not work
    #     '''fetches the main categories from the instance and given a list of group names, checks which one is the "Hauptkategorie"'''
    #     response = ckan_manager.post('/api/3/action/group_list?all_fields=True')
    #     main_categories = []
    #     if response['success']:
    #         groups = response['result']
    #         for group in groups:
    #             title = group['title']
    #             if title == 'Hauptkategorien': 
    #                 main_categories.append(title)

    #     for f_group in fetched_groups:
    #         if f_group in main_categories:
    #             return f_group
    #     return ''

    def find_start_row(self):
        last_row = 0
        for i, row in enumerate(self.ws.iter_rows(values_only=True), start=1):
            if any(cell not in (None, "") for cell in row): 
                last_row = i  
        return last_row + 1 
    
    def write_catalog_to_excel(self, catalog_data):
        # TODO: move that outside the ExcelHandler class
        start_row = self.find_start_row()
        ws = self.ws
        # write catalog data into the xlsx file, first column is empty!
        ws.cell(row=start_row, column=2).value = catalog_data['title']
        ws.cell(row=start_row, column=3).value = catalog_data['notes']  # Beschreibung
        # ws.cell(row=start_row, column=3).value = #Sichtbarkeit
        ws.cell(row=start_row, column=5).value = ';'.join([tag['name'] for tag in catalog_data['tags']])
        ws.cell(row=start_row, column=6).value = catalog_data.get('license_title', '')  # Lizenz  
        # there could be multiple authors/maintainers but we only have room for one author
        # also if the field contains "" script would crash w/o proper handling
        def safe_json_loads(data, default):
            try:
                return json.loads(data)
            except (json.JSONDecodeError, TypeError):
                return default

        author_data = safe_json_loads(catalog_data.get("author", "[]"), [])
        maintainer_data = safe_json_loads(catalog_data.get("maintainer", "[]"), [])

        ws.cell(row=start_row, column=7).value = author_data[0].get('author', '') if author_data else ''  # author name
        ws.cell(row=start_row, column=8).value = author_data[0].get('author_email', '') if author_data else ''  # author email

        ws.cell(row=start_row, column=9).value = maintainer_data[0].get('maintainer', '') if maintainer_data else ''  # maintainer name
        ws.cell(row=start_row, column=10).value = maintainer_data[0].get('Maintainer Email', '') if maintainer_data else ''  # maintainer email
        ws.cell(row=start_row, column=11).value = maintainer_data[0].get('phone', '') if maintainer_data else ''  # phone
        ws.cell(row=start_row, column=12).value = catalog_data.get('organization', {}).get('title', '') # " organisation"

        #groups - list of dicts
        # first get main category
        group_titles = [dic['title'] for dic in catalog_data.get("groups", {})]

        
        #main_cat = self.check_hauptkategorien(group_titles)
        # if main_cat != '': 
        #     group_titles.remove(main_cat)
        #     ws.cell(row=start_row, column=13).value = main_cat # Hauptkategorie
        # else: # randomly take Hauptkategorie since distinguishing between hauptkategorie and themen not possible rn -> Issue #5
        #     ws.cell(row=start_row, column=13).value = group_titles[0]
        #     group_titles.remove(group_titles[0])

        # for i, group in enumerate(group_titles):
        #     if 14+i == 17: # we only have 3 columns that can be filled
        #         break
        #     ws.cell(row=start_row, column=14+i).value = group # Thema

        # write groups separated by semicol in one cell like tags
        ws.cell(row=start_row, column=13).value = ";".join(group_titles)

        language_map = {
        'de': 'Deutsch',
        'en': 'English'
        }

        access_rights_map = {
        # TODO: complete mapping for other possible options e.g. Nur ausgewählte Benutzer
        'public': 'Öffentlich',
        'private': 'Registrierte Benutzer'
        }

        ws.cell(row=start_row, column=17).value = language_map.get(catalog_data.get('language'), '')
        ws.cell(row=start_row, column=18).value = catalog_data.get('version', '')
        ws.cell(row=start_row, column=19).value = catalog_data.get('begin_collection_date', '')# Gültigkeitsdatum Start
        ws.cell(row=start_row, column=20).value = catalog_data.get('end_collection_date', '')# Gültigkeitsdatum Ende

        # resources, 5 columns with 1 empty column in between resources, in total 4 placeholders in excel file 
        for j, resource in enumerate(catalog_data.get("resources", {})):
            ws.cell(row=start_row, column=22+j*6).value = resource["url"]
            ws.cell(row=start_row, column=23+j*6).value = resource["name"]
            ws.cell(row=start_row, column=24+j*6).value = resource["description"]
            ws.cell(row=start_row, column=25+j*6).value = resource["format"]

            restricted = resource["restricted"]

        if restricted:
            try:
                if isinstance(restricted, str):
                    restricted_data = json.loads(restricted)
                elif isinstance(restricted, dict):
                    restricted_data = restricted
                else:
                    restricted_data = {}

                level = restricted_data.get("level", "")
                ws.cell(row=start_row, column=26 + j * 6).value = access_rights_map.get(level, "")
            except (json.JSONDecodeError, TypeError):
                ws.cell(row=start_row, column=26 + j * 6).value = ""
        else:
            ws.cell(row=start_row, column=26 + j * 6).value = ""
                                

        self.wb.save(self.path)

def save_config(config_file='config_write.ini'):
    config = configparser.ConfigParser()
    if os.path.exists(config_file):
        config.read(config_file)
        api_key = config['DEFAULT'].get('API_KEY')
        instance_url = config['DEFAULT'].get('INSTANCE_URL')
        path = config['DEFAULT'].get('excel_file_path')
    else:
        api_key = input("Enter the API key: ")
        while True:
            instance_url = input("Enter the instance url (e.g. http://192.168.92.1:5000): ")
            if is_valid_url(instance_url):
                break
        path = input("Please enter the Excel file path: ").strip()
        if not os.path.exists(path):
            print("The file path does not exist.")
            exit(0)
        config['DEFAULT'] = {'API_KEY': api_key, 'INSTANCE_URL': instance_url, 'EXCEL_FILE_PATH': path}
        with open(config_file, 'w') as configfile:
            config.write(configfile)
    return api_key, instance_url, path

def is_valid_url(url):
    pattern = re.compile(r"^(http|https)://(?:[0-9]{1,3}\.){3}[0-9]{1,3}:[0-9]{1,5}$")
    if pattern.match(url):
        protocol, rest = url.split("://")
        ip, port = rest.split(":")
        parts = ip.split(".")
        if all(0 <= int(part) <= 255 for part in parts):
            if 0 <= int(port) <= 65535:
                return True
    return False

def fetch_and_write_catalog(catalog_name, excel_handler, ckan_manager):
    catalog_data = ckan_manager.fetch_package_data(catalog_name)
    
    excel_handler.write_catalog_to_excel(catalog_data['result'])
    print(f"Successfully wrote catalog '{catalog_data['result']['title']}' to .xlsx file.")

if __name__ == "__main__":
    api_key, api_url, path = save_config()
    ckan_manager = CKANManager(api_url, api_key)
    excel_handler = ExcelHandler(path)

    export_all = input("Do you want to export all catalogs? (yes/no): ").strip().lower()

    if export_all == "yes":
        all_catalogs = ckan_manager.get("/api/3/action/package_list").get("result", [])
        if not all_catalogs:
            print("No catalogs found on the CKAN instance.")
            exit(0)
        print(f"Found {len(all_catalogs)} catalogs. Starting export...")
        for catalog_name in all_catalogs:
            fetch_and_write_catalog(catalog_name, excel_handler, ckan_manager)
        print("Successfully exported all catalogs")
    else:
        catalog_name = input("Enter the CKAN catalog name (id): ").strip()
        fetch_and_write_catalog(catalog_name, excel_handler, ckan_manager)


