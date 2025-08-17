import requests
import json
import openpyxl
import os
import configparser
import re
from datetime import datetime
from ckan_manager import CKANManager
    
class ExcelHandler:
    '''Handles interaction with the Excel file'''
    def __init__(self, path):
        self.path = path
        try:
            if os.path.exists(self.path):
                self.wb = openpyxl.load_workbook(self.path)
            else:
                self.wb = openpyxl.Workbook()  # Create new workbook
                # Removed generate_test_file call
                self.wb.save(self.path)
        except Exception as e:
            print(f"Error with workbook: {str(e)}")
            return

        # Print available sheets for debugging
        print(f"Available sheets: {self.wb.sheetnames}")
        self.ws = self.wb.active  # Use active sheet instead of first sheet

    def find_start_row(self, target, ws_index=0):
        '''Find the start row based on a target value.'''
        print(f"Searching for '{target}' in sheet {self.wb.sheetnames[ws_index]}")
        for i, row in enumerate(self.wb.worksheets[ws_index].iter_rows(values_only=True), start=1):
            print(f"Row {i}: {row}")
            if any(cell == target for cell in row if cell is not None):
                return i
        print(f"Warning: Could not find '{target}' in the worksheet")
        return 1  # Return 1 if target not found, assuming headers are in first row

    def extract_data(self, start_row, ws_index=0):
        '''Extract data starting from a specific row.'''
        if not isinstance(start_row, int):
            print(f"Invalid start_row: {start_row}")
            return []

        data = []
        worksheet = self.wb.worksheets[ws_index]
        print(f"Extracting data from sheet '{worksheet.title}' starting at row {start_row}")
        
        for row in worksheet.iter_rows(min_row=start_row, values_only=True):
            if all(cell is None for cell in row):
                continue  # Skip empty rows
            data.append(row)
        
        if not data:
            print("Warning: No data found in specified range")
        return data

class MetadataManager:
    def __init__(self, main_config_path):
        # add main config
        with open(main_config_path, 'r', encoding='utf-8') as f:
            self.main_config = json.load(f)
        self.schema_mappings = self.main_config.get("schema_mappings", {})
        self.loaded_templates = {}  
        self.organisations = [
            {"title": "Technische Universität München (TUM)", "name": "technische-universitat-munchen"},
            {"title": "Lehrstuhl für Geoinformatik", "name": "lehrstuhl-fur-geoinformatik"},
            {"title": "Bayerische Vermessungsverwaltung", "name": "bayerische-vermessungsverwaltung"},
            {"title": "Bayern Innovativ", "name": "bayern-innovativ"},
            {"title": "Bayerische Landesanstalt für Wald und Forstwirtschaft (LWF)", "name": "bayerische-landesanstalt-wald-forstwirtschaft"},
            {"title": "Bayerische Staatsregierung", "name": "bayerische-staatsregierung"},
            {"title": "Bayerisches Landesamt für Denkmalpflege", "name": "bayerisches-landesamt-landesamt-denkmalpflege"},
            {"title": "Bayerisches Landesamt für Statistik und Datenverarbeitung", "name": "bayerisches-landesamt-statistik-datenverarbeitung"},
            {"title": "Bayerisches Landesamt für Umwelt", "name": "bayerisches-landesamt-umwelt"},
            {"title": "Bayerisches Staatsministerium der Finanzen und für Heimat", "name": "bayerisches-staatsministerium-finanzen-heimat"},
            {"title": "Bayerisches Staatsministerium für Ernährung, Landwirtschaft und Forsten", "name": "bayerisches-staatsministerium-ernaehrung-landwirtschaft-forsten"},
            {"title": "Bayerisches Staatsministerium für Gesundheit und Pflege", "name": "bayerisches-staatsministerium-gesundheit-pflege"},
            {"title": "Bayerisches Staatsministerium für Unterricht und Kultus", "name": "bayerisches-staatsministerium-unterricht-kultus"},
            {"title": "Bayerisches Staatsministerium für Wirtschaft, Landesentwicklung und Energie", "name": "bayerisches-staatsministerium-wirtschaft-landesentwicklung-energie"},
            {"title": "Bayerisches Staatsministerium für Wohnen, Bau und Verkehr", "name": "bayerisches-staatsministerium-wohnen-bau-verkehr"},
            {"title": "Technologieanbieter", "name": "technologieanbieter"},
            {"title": "Softwareanbieter", "name": "softwareanbieter"},
            {"title": "Hardwareentwickler", "name": "hardwareentwickler"}
        ]

        self.groups = [
            {"title": "Hauptkategorien", "name": "main-categories"},
            {"title": "Datensatz und Dokumente", "name": "dataset"},
            {"title": "Online-Dienst", "name": "online-service"},
            {"title": "Projekt", "name": "project"},
            {"title": "Software", "name": "software"},
            {"title": "Online-Anwendung", "name": "online-application"},
            {"title": "Methode", "name": "method"},
            {"title": "Gerät / Ding", "name": "device"},
            {"title": "Geoobjekt", "name": "geoobject"},
            {"title": "Digitaler Zwilling", "name": "digital-twin"},
            
            {"title": "Themen", "name": "topics"},
            {"title": "Verwaltung", "name": "administration"},
            {"title": "Stadtplanung", "name": "urban-planning"},
            {"title": "Umwelt", "name": "environment"},
            {"title": "Gesundheit", "name": "health"},
            {"title": "Energie", "name": "energy"},
            {"title": "Informations-Technologie", "name": "information-technology"},
            {"title": "Tourismus & Freizeit", "name": "tourism"},
            {"title": "Wohnen", "name": "living"},
            {"title": "Bildung", "name": "education"},
            {"title": "Handel", "name": "trade"},
            {"title": "Bauen", "name": "construction"},
            {"title": "Kultur", "name": "culture"},
            {"title": "Mobilität", "name": "mobility"},
            {"title": "Landwirtschaft", "name": "agriculture"},
            {"title": "Gewerbe / Handwerk", "name": "craft"},
            {"title": "Arbeiten", "name": "work"}
        ]
        self.access_rights = {
            'Öffentlich': 'public',
            'Registrierte Benutzer': 'registered',
            'Mitglieder der selben Organisation': 'same_organization',
            'Nur ausgewählte Benutzer': 'only_allowed_users'
        }

    def get_template_for_schema(self, schema_type):
        # if loaded, return
        if schema_type in self.loaded_templates:
            return self.loaded_templates[schema_type]
        # otherwise load
        template_path = self.schema_mappings.get(schema_type)
        if not template_path or not os.path.exists(template_path):
            raise Exception(f"Template for schema '{schema_type}' not found: {template_path}")
        with open(template_path, 'r', encoding='utf-8') as f:
            template = json.load(f)
        self.loaded_templates[schema_type] = template
        return template

    def validate_and_construct_package(self, row_data, schema_type):
        template = self.get_template_for_schema(schema_type)
        field_mappings = template.get("field_mappings", {})
        package_data = {}
        for excel_col, ckan_field in field_mappings.items():
            package_data[ckan_field] = row_data.get(excel_col, "")

        # 2. tags 
        if "tag_string" in package_data:
            package_data["tags"] = convert_string_to_tags(package_data["tag_string"])
            # Do not pop tag_string, CKAN needs it

        # 3. composite fields
        def build_composite(prefix, fields):
            obj = {}
            for f in fields:
                key = f"{prefix}__{f}"
                if key in package_data:
                    obj[f] = package_data[key]
            return obj

        composite_fields = {
            "author": ["author_name", "author_email", "role"],
            "maintainer": ["maintainer_name", "maintainer_email", "phone", "role"]
        }
        for comp, subfields in composite_fields.items():
            comp_obj = build_composite(comp, subfields)
            if any(comp_obj.values()):
                package_data[comp] = json.dumps([comp_obj], ensure_ascii=False)
            else:
                package_data[comp] = json.dumps([], ensure_ascii=False)
        # Do not pop/del author__author_name etc.

        # 4. group
        group_fields = [k for k in field_mappings.values() if k in ["main_category", "theme", "group"]]
        groups = []
        def get_group_name(value):
            for g in self.groups:
                if value == g['title'] or value == g['name']:
                    return g['name']
            return value
        for field in group_fields:
            val = package_data.get(field, "")
            if val:
                name = get_group_name(val)
                if name and name not in [g['name'] for g in groups]:
                    groups.append({'name': name})
        package_data['groups'] = groups
        # pop intermediate fields
        for field in group_fields:
            if field in package_data:
                del package_data[field]

        # 5. license_id
        if "license_id" in package_data and hasattr(self, "licenses"):
            lic_title = row_data.get("License", "")
            if lic_title in self.licenses:
                package_data["license_id"] = self.licenses[lic_title]
            else:
                package_data["license_id"] = 'notspecified'

        # 6. private field handling
        if "private" in package_data:
            sichtbarkeit = str(package_data.get("private", "")).strip()
            # Fix visibility logic
            package_data["private"] = sichtbarkeit != "Öffentlich"
        else:
            # Default to public
            package_data["private"] = False

        # Ensure state is active
        package_data["state"] = "active"

        # 7. resources
        url = row_data.get('Datei/ Link')
        if url:
            resource = {
                'url': url,
                'name': row_data.get('Name', ''),
                'description': row_data.get('Beschreibung', ''),
                'format': row_data.get('Format', ''),
                'restricted': {
                    'level': self.access_rights.get(row_data.get('zugriffsrechte', ''), 'public'),
                    'allowed_users': ""
                }
            }
            resource['restricted_level'] = self.access_rights.get(row_data.get('zugriffsrechte', ''), 'public')
            package_data['resources'] = [resource]
        else:
            package_data['resources'] = []

        # 8. name 
        if "title" in package_data:
            lower_case = package_data["title"].lower()
            alphanumeric = re.sub(r'[^a-z0-9]', '', lower_case)
            package_data['name'] = alphanumeric

        # 9. licence_agreement -> list
        if "licence_agreement" in package_data:
            val = package_data["licence_agreement"]
            if isinstance(val, str) and val:
                package_data["licence_agreement"] = [val]
            elif not val:
                package_data["licence_agreement"] = []

        # 10. date fields
        for date_field in ["begin_collection_date", "end_collection_date"]:
            if date_field in package_data and package_data[date_field]:
                val = package_data[date_field]
                if isinstance(val, datetime):
                    package_data[date_field] = val.strftime("%Y-%m-%d")
                else:
                    package_data[date_field] = str(val)

        # 11. spatial field validation
        if "spatial" in package_data:
            spatial_val = package_data.get("spatial", "")
            if spatial_val and spatial_val.strip():
                # Try to validate as JSON
                try:
                    if isinstance(spatial_val, str):
                        # Try to parse as JSON to validate
                        json.loads(spatial_val)
                    # If it's valid JSON, keep it as is
                    package_data["spatial"] = spatial_val
                except (json.JSONDecodeError, TypeError):
                    # If invalid JSON, remove the spatial field to prevent API errors
                    print(f"WARNING: Invalid spatial data detected, removing spatial field: {spatial_val}")
                    del package_data["spatial"]
            else:
                # If empty or None, remove the field
                del package_data["spatial"]
        
        # 12. handel multi values
        if schema_type.lower() == "onlineservice":
            for excel_col, ckan_field in field_mappings.items():
                if "supported_method" in ckan_field.lower():
                    raw_val = row_data.get(excel_col, "")
                    if isinstance(raw_val, str) and ";" in raw_val:
                        methods = [m.strip() for m in raw_val.split(";") if m.strip()]
                        package_data[ckan_field] = methods
                    elif isinstance(raw_val, str) and raw_val.strip():
                        package_data[ckan_field] = [raw_val.strip()]
                    else:
                        package_data[ckan_field] = []
        
        # Handle digitaltwin multi-value fields
        elif schema_type.lower() == "digitaltwin":
            for excel_col, ckan_field in field_mappings.items():
                if ckan_field == "twin_capabilities":
                    raw_val = row_data.get(excel_col, "")
                    if isinstance(raw_val, str) and ";" in raw_val:
                        capabilities = [c.strip() for c in raw_val.split(";") if c.strip()]
                        package_data[ckan_field] = capabilities
                    elif isinstance(raw_val, str) and raw_val.strip():
                        package_data[ckan_field] = [raw_val.strip()]
                    else:
                        # 如果为空，设置为空列表而不是空字符串
                        package_data[ckan_field] = []

        org_title = package_data.get("owner_org", "")
        org_name = None
        for org in self.organisations:
            if org_title == org["title"] or org_title == org["name"]:
                org_name = org["name"]
                break
        if org_name:
            package_data["owner_org"] = org_name
        else:
            raise Exception(f"Organization '{org_title}' not found in available organizations!")

        package_data["type"] = schema_type  

        return package_data


# helper functions
def convert_string_to_tags(input_string):
    if not input_string:  # Handle None or empty string
        return []
    words = [word.strip() for word in str(input_string).split(';') if word.strip()]
    return [{'name': word} for word in words]

def load_config(config_file='config.ini'):
    config = configparser.ConfigParser()
    config.read(config_file)
    api_key = config.get('DEFAULT', 'api_key', fallback=None)
    instance_url = config.get('DEFAULT', 'instance_url', fallback=None)
    path = config.get('DEFAULT', 'excel_file_path', fallback=None)
    schema_config = config.get('DEFAULT', 'schema_config', fallback=None)
    return api_key, instance_url, path, schema_config

def compare_mapped_fields(sheet_data, ckan_data, field_mappings):
    """Compare all mapped fields between sheet data and CKAN data. Return True if different, False if same."""
    for excel_col, ckan_field in field_mappings.items():
        sheet_val = sheet_data.get(excel_col, "")
        ckan_val = ckan_data.get(ckan_field, "")
        # Normalize for comparison
        if isinstance(sheet_val, (list, tuple)):
            sheet_val = list(sheet_val)
        if isinstance(ckan_val, (list, tuple)):
            ckan_val = list(ckan_val)
        if str(sheet_val).strip() != str(ckan_val).strip():
            return True  # Difference found
    return False

if __name__ == "__main__":
    api_key, api_url, path, schema_config = load_config()
    ckan_manager = CKANManager(api_url, api_key)
    # Do NOT create or delete organizations/groups
    # cleanup_script_created_items(ckan_manager)
    # create_organizations_and_groups(ckan_manager, metadata_manager)
    metadata_manager = MetadataManager("schema_templates/schema_config.json")
    excel_handler = ExcelHandler(path)

    # Define schema sheets to process (excluding 'Data Mapping')
    schema_sheets = ['dataset', 'device', 'digitaltwin', 'geoobject', 
                    'onlineapplication', 'onlineservice', 'project', 'software']

    print("\nProcessing schema sheets...")
    for sheet_name in schema_sheets:
        try:
            sheet_index = excel_handler.wb.sheetnames.index(sheet_name)
            worksheet = excel_handler.wb.worksheets[sheet_index]
            headers = list(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if all(cell is None or cell == '' for cell in row):
                    continue
                try:
                    row_list = list(row)
                    row_data = dict(zip(headers, row_list))
                    package_data = metadata_manager.validate_and_construct_package(row_data, sheet_name)
                    catalog_title = package_data.get('title', '[No Title]')
                    dataset_name = package_data.get('name')
                    field_mappings = metadata_manager.get_template_for_schema(sheet_name).get('field_mappings', {})
                    
                    try:
                        # Check if dataset exists
                        show_resp = ckan_manager.post('/api/3/action/package_show', {"id": dataset_name})
                        if show_resp.get("success"):
                            existing = show_resp["result"]
                            # Compare fields
                            if compare_mapped_fields(row_data, existing, field_mappings):
                                # Update if different
                                update_data = existing.copy()
                                update_data.update(package_data)
                                update_data["id"] = existing["id"]
                                # Remove fields that shouldn't be updated
                                for key in ["revision_id", "metadata_created", "metadata_modified", "creator_user_id"]:
                                    update_data.pop(key, None)
                                
                                print(f"Updating existing dataset: {catalog_title}")
                                try:
                                    update_resp = ckan_manager.post('/api/3/action/package_update', update_data)
                                    if update_resp.get("success"):
                                        print(f"Successfully UPDATED: {catalog_title}")
                                    else:
                                        error_details = update_resp.get('error', {})
                                        if 'spatial' in str(error_details):
                                            # Try again without spatial field
                                            print(f"Spatial field error detected, retrying without spatial data...")
                                            update_data_no_spatial = update_data.copy()
                                            update_data_no_spatial.pop('spatial', None)
                                            retry_resp = ckan_manager.post('/api/3/action/package_update', update_data_no_spatial)
                                            if retry_resp.get("success"):
                                                print(f"Successfully UPDATED (without spatial): {catalog_title}")
                                            else:
                                                print(f"Failed to update: {catalog_title} - {retry_resp}")
                                        else:
                                            print(f"Failed to update: {catalog_title} - {update_resp}")
                                except Exception as update_error:
                                    print(f"Update error for {catalog_title}: {str(update_error)}")
                            else:
                                print(f"No updates needed for: {catalog_title}")
                        else:
                            # Dataset exists but API call failed for other reasons
                            print(f"Failed to retrieve dataset {catalog_title}: {show_resp}")
                    except Exception as e:
                        error_msg = str(e)
                        # Check if it's a "Not found" error (dataset doesn't exist)
                        if "404" in error_msg or "Not found" in error_msg or "Not Found" in error_msg:
                            # Create new dataset since it doesn't exist
                            print(f"Dataset {catalog_title} not found, creating new dataset...")
                            try:
                                response = ckan_manager.post('/api/3/action/package_create', package_data)
                                if response.get('success'):
                                    print(f"Successfully CREATED: {catalog_title}")
                                else:
                                    print(f"Failed to create: {catalog_title} - {response}")
                            except Exception as create_error:
                                print(f"Create error for {catalog_title}: {str(create_error)}")
                        else:
                            # Other error, not related to dataset existence
                            print(f"Error processing dataset {catalog_title}: {error_msg}")
                        continue
                except Exception as e:
                    print(f"FAILED: [Unknown Title] - {str(e)}")
                    continue
        except ValueError as e:
            continue
        except Exception as e:
            continue
    print("\nProcessing complete!")